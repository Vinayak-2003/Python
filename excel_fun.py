import openpyxl as pyxl
from openpyxl import Workbook

path = r"C:\\Users\\VinayakKanchan\\Desktop\\Programming\\Python\\excel\\task1.xlsx"
wb = pyxl.load_workbook(path)
sheet = wb.active
new_file = Workbook()
new_sheet = new_file.active

# for i in range (1, sheet.max_column+1):
#     cell_obj = sheet.cell(row=1, column=i)
#     if cell_obj.value == 'Schedule Id':
#         for j in range(1, sheet.max_row):
#             r_value = sheet.cell(row=j+1, column=i)
#             r_value.value="NEW"

for i in range (1, sheet.max_column+1):
    cell_obj = sheet.cell(row=1, column=i)
    if cell_obj.value == 'Data Id':
        for j in range(1, sheet.max_row):
            r_value = sheet.cell(row=j, column=i)
            val = list(r_value.value)
            new_sheet.append(val)
            
            
    
new_file.save(r"C:\\Users\\VinayakKanchan\\Desktop\\Programming\\Python\\excel\\task1_NEW.xlsx")