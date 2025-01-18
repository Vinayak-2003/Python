import openpyxl as pyxl

# location
path = r"C:\\Users\\VinayakKanchan\\Downloads\\task1.xlsx"

# an object is created to open the workbook
wb_obj = pyxl.load_workbook(path)

# get workbook active
sheet_obj = wb_obj.active

# cell_obj = sheet_obj.cell(row=1, column=1)

# print(cell_obj.value)

# ------------------------------------------------
row = sheet_obj.max_row
col = sheet_obj.max_column

print("TOtal row = ", row)
print("Column total = ", col)

for i in range (1, col+1):
    cell_obj = sheet_obj.cell(row=1, column=i)
    if cell_obj.value == 'Schedule Id':
        for j in range(1, row):
            r_value = sheet_obj.cell(row=j+1, column=i)
            r_value.value="New"
            
      
# for i in range(2, row+1):
#     cell = f'B{i}'
#     value = 'NEW'
#     sheet_obj[cell] = value
                
wb_obj.save(r"C:\\Users\\VinayakKanchan\\Downloads\\task1_NEW.xlsx")
    
    